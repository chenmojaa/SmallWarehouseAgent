"""Parse .xlsx -> Markdown text via openpyxl, preserving merged cells."""
from __future__ import annotations

import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _md_escape(s: str) -> str:
    # Markdown table cells only need to escape the pipe and newlines.
    return s.replace("|", "\\|").replace("\n", " ").replace("\r", " ")


def _expand_sheet_to_grid(ws) -> tuple[list[list[str]], list[str]]:
    """Return (grid, col_widths_hint) where:
      - grid is a list of rows (each a list of cell strings).
      - Cells inside a merged range but NOT the top-left are returned as "".
    """
    # Determine grid extent (max_row / max_col from iter is unreliable for
    # merged-only content), so we read every cell that has a value OR is part
    # of a merged range.
    max_r = ws.max_row or 0
    max_c = ws.max_column or 0
    for mr in ws.merged_cells.ranges:
        if mr.max_row > max_r:
            max_r = mr.max_row
        if mr.max_col > max_c:
            max_c = mr.max_col
    if max_r == 0 or max_c == 0:
        return [], []

    # Seed grid with empty strings.
    grid: list[list[str]] = [["" for _ in range(max_c)] for _ in range(max_r)]

    # Build a set of "non-anchor" cells inside merged ranges so we skip them.
    skip: set[tuple[int, int]] = set()
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    skip.add((r, c))

    for row in ws.iter_rows(min_row=1, max_row=max_r,
                            min_col=1, max_col=max_c,
                            values_only=False):
        for cell in row:
            r, c = cell.row, cell.column
            if (r, c) in skip:
                continue
            grid[r - 1][c - 1] = _cell(cell.value)

    col_widths = [get_column_letter(c + 1) for c in range(max_c)]
    return grid, col_widths


def _grid_to_markdown(grid: list[list[str]]) -> str:
    if not grid:
        return ""
    n_cols = max((len(r) for r in grid), default=0)
    # Pad short rows so columns line up.
    rows = [r + [""] * (n_cols - len(r)) for r in grid]

    # Drop fully empty trailing rows.
    while rows and not any(c.strip() for c in rows[-1]):
        rows.pop()

    if not rows:
        return ""

    # Compute column widths for nicer alignment (not strictly required).
    widths = [0] * n_cols
    for r in rows:
        for i, c in enumerate(r):
            widths[i] = max(widths[i], len(_md_escape(c)))

    def fmt_row(r: list[str]) -> str:
        cells = [_md_escape(c).ljust(widths[i]) for i, c in enumerate(r)]
        return "| " + " | ".join(cells) + " |"

    header = rows[0]
    sep = "| " + " | ".join("-" * widths[i] for i in range(n_cols)) + " |"
    body = "\n".join(fmt_row(r) for r in rows[1:])
    out = fmt_row(header) + "\n" + sep
    if body:
        out += "\n" + body
    return out


def parse_xlsx(path: str) -> dict:
    """Extract all sheet text as Markdown tables. Merged cells keep their anchor value."""
    if not os.path.isfile(path):
        raise FileNotFoundError(path)
    wb = load_workbook(path, data_only=True, read_only=False)
    parts = []
    for sheet in wb.worksheets:
        grid, _cols = _expand_sheet_to_grid(sheet)
        if not grid:
            continue
        md = _grid_to_markdown(grid)
        if md:
            parts.append(f"### Sheet: {sheet.title}\n\n{md}")
    wb.close()

    content = "\n\n".join(parts).strip()
    title = os.path.splitext(os.path.basename(path))[0] or "Untitled Spreadsheet"
    if not content:
        content = "(empty spreadsheet)"
    return {"title": title[:200], "content": content}
