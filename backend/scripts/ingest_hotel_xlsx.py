"""Ingest hotel price xlsx into the knowledge base.

Bypasses the 50MB upload limit by reading the file from disk directly.
Streams rows in read_only mode (low memory) and chunks per-row so each
hotel/date/price combination becomes its own retrieval unit.

Usage (from backend/):
    .\.venv\Scripts\python.exe scripts\ingest_hotel_xlsx.py ^
        --path "C:/Users/Administrator/Desktop/jd_spider_tuniu_hotel_price.xlsx" ^
        --max-rows-per-sheet 0 ^
        --api-key %KEY% --base-url "https://api.minimax.chat/v1"

max-rows-per-sheet 0 = unlimited (full ingest).
"""
import argparse, os, sys, time
from datetime import datetime, date

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.abspath(os.path.join(HERE, "..")))

import openpyxl

from app.storage.db import get_engine, Note
from app.storage.vector import add_chunks
from app.storage import llm_config_store as _lcs
from app.config import settings
from app.embeddings.factory import embed_texts
from sqlmodel import Session as SqlSession


COLS = ["id", "city", "hotel_name", "price", "room_type",
        "check_in_date", "check_out_date", "business", "crawled_at", "hotel_id"]


def _fmt(v):
    if v is None or v == "":
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def row_to_text(row):
    parts = []
    if row.get("hotel_name"):
        parts.append("\u9152\u5e97 " + row["hotel_name"])
    if row.get("city"):
        parts.append("\u57ce\u5e02 " + row["city"])
    if row.get("room_type"):
        parts.append("\u623f\u578b " + row["room_type"])
    if row.get("check_in_date"):
        parts.append("\u5165\u4f4f " + row["check_in_date"])
    if row.get("check_out_date"):
        parts.append("\u79bb\u5e97 " + row["check_out_date"])
    if row.get("price") not in (None, ""):
        parts.append("\u4ef7\u683c " + row["price"] + "\u5143")
    if row.get("business"):
        parts.append("\u5546\u5708 " + row["business"])
    if row.get("hotel_id"):
        parts.append("hotel_id=" + row["hotel_id"])
    return "\uff0c".join(parts) or ("hotel_id=" + row.get("hotel_id", ""))


def ingest_sheet(ws, *, max_rows, api_key, base_url):
    """Ingest one worksheet. max_rows=0 means unlimited (full sheet)."""
    rows_iter = ws.iter_rows(min_row=2, values_only=True)
    chunks = []
    row_dicts = []
    skipped = 0
    report_every = 5000
    started = time.time()
    for raw in rows_iter:
        if max_rows and len(chunks) >= max_rows:
            break
        rd = {COLS[i]: (_fmt(raw[i]) if i < len(raw) else "") for i in range(len(COLS))}
        if not rd["hotel_name"] and not rd["hotel_id"]:
            skipped += 1
            continue
        chunks.append(row_to_text(rd))
        row_dicts.append(rd)
        if len(chunks) % report_every == 0:
            rate = len(chunks) / max(1, time.time() - started)
            print("    " + str(len(chunks)) + " rows (" + str(round(rate,1)) + " rows/s)", flush=True)
    if not chunks:
        return None, 0, 0
    print("  embedding " + str(len(chunks)) + " rows in batches ...", flush=True)
    t0 = time.time()
    embs = embed_texts(chunks, api_key=api_key, base_url=base_url)
    print("  embedded in " + str(round(time.time()-t0,1)) + "s, dim=" + str(len(embs[0]) if embs else 0), flush=True)
    note_id = "n_" + os.urandom(6).hex()
    os.makedirs(settings.notes_dir, exist_ok=True)
    cp = os.path.join(settings.notes_dir, note_id + ".md")
    with open(cp, "w", encoding="utf-8") as f:
        f.write("# " + ws.title + "\n\n")
        for txt in chunks:
            f.write("- " + txt + "\n")
    note = Note(
        id=note_id, title=("\u9152\u5e97\u4ef7\u683c_" + ws.title)[:500],
        source_type="xlsx", source_url=None, content_path=cp,
        word_count=sum(len(c) for c in chunks),
        chunk_count=len(chunks), embedded=True,
        summary=chunks[0][:200], created_at=datetime.utcnow(),
    )
    engine = get_engine()
    with SqlSession(engine) as s:
        s.add(note); s.commit(); s.refresh(note)
    add_chunks(note_id, chunks, embs)
    return note, len(chunks), skipped


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--path", required=True)
    ap.add_argument("--max-rows-per-sheet", type=int, default=2000,
                    help="0 = unlimited (full sheet)")
    ap.add_argument("--api-key", default=None)
    ap.add_argument("--base-url", default=None)
    ap.add_argument("--only-sheets", default=None)
    args = ap.parse_args()
    api_key = args.api_key or _lcs.get_api_key()
    base_url = args.base_url or _lcs.get_base_url() or None
    print("opening " + args.path, flush=True)
    wb = openpyxl.load_workbook(args.path, read_only=True, data_only=True)
    targets = wb.sheetnames
    if args.only_sheets:
        wanted = set(s.strip() for s in args.only_sheets.split(","))
        targets = [s for s in targets if s in wanted]
    for name in targets:
        ws = wb[name]
        cap = "unlimited" if not args.max_rows_per_sheet else str(args.max_rows_per_sheet)
        print("sheet: " + name + " (rows=" + str(ws.max_row) + ", cap=" + cap + ")", flush=True)
        t0 = time.time()
        note, n, skipped = ingest_sheet(ws, max_rows=args.max_rows_per_sheet,
            api_key=api_key, base_url=base_url)
        if note is None:
            print("  (empty)"); continue
        print("  -> note_id=" + note.id + " chunks=" + str(n) + " skipped=" + str(skipped) +
              " in " + str(round(time.time()-t0,1)) + "s", flush=True)
    wb.close()
    print("done")


if __name__ == "__main__":
    main()